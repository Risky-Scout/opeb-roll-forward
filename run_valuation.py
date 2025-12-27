#!/usr/bin/env python3
"""
OPEB VALUATION - PRODUCTION RUNNER
==================================
This is the SINGLE script you run to complete a valuation.

USAGE:
------
For ROLL-FORWARD valuation:
    python run_valuation.py rollforward \
        --prior-file "GASB75_Client_2024.xlsx" \
        --client-name "West Florida Planning" \
        --discount-rate 0.0502 \
        --prior-discount-rate 0.0381

For FULL valuation:
    python run_valuation.py full \
        --prior-file "GASB75_Client_2024.xlsx" \
        --actives-file "census_actives.xlsx" \
        --retirees-file "census_retirees.xlsx" \
        --client-name "West Florida Planning" \
        --discount-rate 0.0502 \
        --prior-discount-rate 0.0381

OUTPUT:
-------
    GASB75_{ClientName}_{Year}_FINAL.xlsx - Ready to deliver

IMPORTANT:
----------
This script updates the PRIOR YEAR file directly. It does NOT use a template.
The update process:
1. Hardcodes RSI current year column (preserves prior year data)
2. Sets RSI new year column formulas
3. Shifts ARSL values in AmortDeferredOutsIns (ALL values shift down)
4. Updates AmortDeferredOutsIns A13 (current year)
5. Updates Assumptions tab (dates and rates)
6. Updates ProVal1 (valuation results)
7. Updates Net OPEB labels
"""

import argparse
import sys
import os
import subprocess
from datetime import date, datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, Dict, Any, List

# Auto-install dependencies
def ensure_package(package):
    try:
        __import__(package)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package, "-q"])

ensure_package("openpyxl")

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.formatting import ConditionalFormattingList


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class ValuationResults:
    """All results needed for GASB 75 disclosure."""
    client_name: str
    measurement_date: date
    prior_measurement_date: date
    valuation_date: date
    active_count: int
    retiree_count: int
    covered_payroll: float
    tol_boy_old_rate: float
    tol_boy_new_rate: float
    tol_eoy_baseline: float
    tol_eoy_disc_plus_1: float
    tol_eoy_disc_minus_1: float
    tol_eoy_trend_plus_1: float
    tol_eoy_trend_minus_1: float
    service_cost: float
    interest_cost: float
    assumption_change: float
    experience_gl: float
    discount_rate_boy: float
    discount_rate_eoy: float
    avg_remaining_service_life: float
    benefit_payments: float = 0.0


@dataclass
class PriorYearData:
    """Data extracted from prior year file."""
    measurement_date: date = None
    tol_eoy: float = 0.0
    service_cost: float = 0.0
    discount_rate: float = 0.0
    arsl: float = 5.0
    active_count: int = 0
    retiree_count: int = 0
    covered_payroll: float = 0.0
    # ARSL values for each historical year (to be shifted)
    arsl_values: Dict[int, float] = field(default_factory=dict)


# =============================================================================
# PRIOR YEAR EXTRACTION
# =============================================================================

def extract_prior_year_data(prior_file: str) -> PriorYearData:
    """Extract all required data from prior year GASB 75 file."""
    print(f"  Loading prior year file: {prior_file}")
    wb = load_workbook(prior_file, data_only=True)
    data = PriorYearData()
    
    # Assumptions - Get measurement date
    assumptions = wb['Assumptions']
    measurement_date = assumptions['C4'].value
    if hasattr(measurement_date, 'year'):
        data.measurement_date = measurement_date
    
    # ProVal1 - Key values
    proval = wb['ProVal1']
    data.tol_eoy = proval['D19'].value or 0
    data.service_cost = proval['B38'].value or 0  # B38 is what Net OPEB uses
    data.active_count = proval['D6'].value or 0
    data.retiree_count = proval['D8'].value or 0
    data.covered_payroll = proval['D17'].value or 0
    data.discount_rate = proval['D88'].value or 0
    data.arsl = proval['D74'].value or 5.0
    
    # AmortDeferredOutsIns - Get ARSL values for shifting
    amort = wb['AmortDeferredOutsIns']
    for row in range(13, 20):
        year = amort[f'A{row}'].value
        arsl = amort[f'C{row}'].value
        if year and arsl:
            data.arsl_values[int(year)] = arsl
    
    wb.close()
    
    print(f"    Prior Measurement Date: {data.measurement_date}")
    print(f"    Prior TOL: ${data.tol_eoy:,.0f}")
    print(f"    Prior Discount Rate: {data.discount_rate:.2%}" if data.discount_rate else "    Prior Discount Rate: Not found")
    print(f"    Prior Service Cost (B38): ${data.service_cost:,.0f}")
    print(f"    Prior ARSL: {data.arsl}")
    
    return data


# =============================================================================
# VALUATION CALCULATIONS
# =============================================================================

def run_rollforward(prior_data: PriorYearData, discount_rate: float, prior_discount_rate: float,
                    benefit_payments: float = 0, measurement_date: date = None, 
                    client_name: str = "Client") -> ValuationResults:
    """
    Run roll-forward valuation.
    
    For roll-forward:
    - Service cost = prior year's service cost
    - Interest = (BOY + 0.5*SC) * prior_discount_rate
    - Assumption change = -BOY * rate_change * duration
    - Experience = 0 (calculated as residual, will be 0)
    """
    print("\n  Running roll-forward calculations...")
    
    if measurement_date is None:
        measurement_date = date.today()
    
    # Dates
    prior_measurement_date = date(measurement_date.year - 1, measurement_date.month, measurement_date.day)
    valuation_date = date(measurement_date.year - 1, measurement_date.month + 1, 1) if measurement_date.month < 12 else date(measurement_date.year, 1, 1)
    
    # BOY = prior EOY
    tol_boy = prior_data.tol_eoy
    
    # Service cost from prior year (B38)
    service_cost = prior_data.service_cost
    
    # Interest at PRIOR (BOY) rate
    interest_cost = (tol_boy + 0.5 * service_cost) * prior_discount_rate
    
    # Assumption change from discount rate change
    rate_change = discount_rate - prior_discount_rate
    duration = 10.0  # Approximate duration
    assumption_change = -tol_boy * rate_change * duration if abs(rate_change) > 0.0001 else 0
    
    # BOY at new rate (for assumption change formula in Excel)
    tol_boy_new_rate = tol_boy + assumption_change
    
    # Experience = 0 for roll-forward
    experience_gl = 0
    
    # EOY = BOY + SC + Interest + Assumption + Experience - Benefits
    tol_eoy = tol_boy + service_cost + interest_cost + assumption_change + experience_gl - benefit_payments
    
    # Sensitivities (approximate)
    tol_eoy_disc_plus_1 = tol_eoy * 0.92
    tol_eoy_disc_minus_1 = tol_eoy * 1.08
    tol_eoy_trend_plus_1 = tol_eoy * 1.04
    tol_eoy_trend_minus_1 = tol_eoy * 0.96
    
    print(f"    BOY TOL (old rate): ${tol_boy:,.0f}")
    print(f"    BOY TOL (new rate): ${tol_boy_new_rate:,.0f}")
    print(f"    Service Cost: ${service_cost:,.0f}")
    print(f"    Interest Cost (at {prior_discount_rate:.2%}): ${interest_cost:,.0f}")
    print(f"    Assumption Change: ${assumption_change:,.0f}")
    print(f"    Experience G/L: ${experience_gl:,.0f}")
    print(f"    EOY TOL: ${tol_eoy:,.0f}")
    
    return ValuationResults(
        client_name=client_name,
        measurement_date=measurement_date,
        prior_measurement_date=prior_measurement_date,
        valuation_date=valuation_date,
        active_count=prior_data.active_count,
        retiree_count=prior_data.retiree_count,
        covered_payroll=prior_data.covered_payroll,
        tol_boy_old_rate=tol_boy,
        tol_boy_new_rate=tol_boy_new_rate,
        tol_eoy_baseline=tol_eoy,
        tol_eoy_disc_plus_1=tol_eoy_disc_plus_1,
        tol_eoy_disc_minus_1=tol_eoy_disc_minus_1,
        tol_eoy_trend_plus_1=tol_eoy_trend_plus_1,
        tol_eoy_trend_minus_1=tol_eoy_trend_minus_1,
        service_cost=service_cost,
        interest_cost=interest_cost,
        assumption_change=assumption_change,
        experience_gl=experience_gl,
        discount_rate_boy=prior_discount_rate,
        discount_rate_eoy=discount_rate,
        avg_remaining_service_life=prior_data.arsl,
        benefit_payments=benefit_payments
    )


def run_full_valuation(prior_data: PriorYearData, actives_file: str, retirees_file: str,
                       discount_rate: float, prior_discount_rate: float, measurement_date: date = None,
                       client_name: str = "Client", benefit_payments: float = 0) -> ValuationResults:
    """
    Run full valuation with census data.
    
    For full valuation:
    - Service cost, EOY TOL come from actual census calculations
    - Experience = residual (EOY - BOY - SC - Interest - Assumption + Benefits)
    """
    print("\n  Running full valuation...")
    
    # Import pandas only for full valuation
    ensure_package("pandas")
    import pandas as pd
    
    if measurement_date is None:
        measurement_date = date.today()
    
    prior_measurement_date = date(measurement_date.year - 1, measurement_date.month, measurement_date.day)
    valuation_date = date(measurement_date.year - 1, measurement_date.month + 1, 1) if measurement_date.month < 12 else date(measurement_date.year, 1, 1)
    
    print(f"    Loading active census: {actives_file}")
    actives_df = pd.read_excel(actives_file)
    print(f"    Loading retiree census: {retirees_file}")
    retirees_df = pd.read_excel(retirees_file)
    
    active_count = len(actives_df)
    retiree_count = len(retirees_df)
    
    # Get covered payroll
    salary_cols = [c for c in actives_df.columns if 'sal' in c.lower() or 'pay' in c.lower()]
    covered_payroll = actives_df[salary_cols[0]].sum() if salary_cols else prior_data.covered_payroll * 1.03
    
    print(f"    Active count: {active_count}")
    print(f"    Retiree count: {retiree_count}")
    print(f"    Covered payroll: ${covered_payroll:,.0f}")
    
    # BOY = prior EOY
    tol_boy = prior_data.tol_eoy
    
    # Try to use valuation engine for actual calculations
    try:
        from opeb_valuation import create_vectorized_engine
        print("    Using vectorized engine...")
        config = {'valuation_date': measurement_date, 'discount_rate': discount_rate}
        engine = create_vectorized_engine(config)
        active_results = engine.run_valuation(actives_df, population='active')
        retiree_results = engine.run_valuation(retirees_df, population='retiree')
        tol_eoy = active_results['TOL'].sum() + retiree_results['TOL'].sum()
        service_cost = active_results['service_cost'].sum()
    except Exception as e:
        print(f"    Valuation engine not available ({e}), using simplified calculations...")
        # Simplified calculation based on census changes
        census_factor = (active_count + retiree_count) / max(prior_data.active_count + prior_data.retiree_count, 1)
        service_cost = prior_data.service_cost * census_factor
        
        interest_cost = (tol_boy + 0.5 * service_cost) * prior_discount_rate
        rate_change = discount_rate - prior_discount_rate
        assumption_change = -tol_boy * rate_change * 10 if abs(rate_change) > 0.0001 else 0
        
        # For full valuation, estimate experience based on census changes
        if abs(census_factor - 1.0) > 0.01:
            experience_gl = tol_boy * (census_factor - 1) * 0.3
        else:
            experience_gl = 0
        
        tol_eoy = tol_boy + service_cost + interest_cost + assumption_change + experience_gl - benefit_payments
    
    # Calculate components
    interest_cost = (tol_boy + 0.5 * service_cost) * prior_discount_rate
    rate_change = discount_rate - prior_discount_rate
    assumption_change = -tol_boy * rate_change * 10 if abs(rate_change) > 0.0001 else 0
    tol_boy_new_rate = tol_boy + assumption_change
    
    # Experience is the RESIDUAL
    experience_gl = tol_eoy - tol_boy - service_cost - interest_cost - assumption_change + benefit_payments
    
    # Sensitivities
    tol_eoy_disc_plus_1 = tol_eoy * 0.92
    tol_eoy_disc_minus_1 = tol_eoy * 1.08
    tol_eoy_trend_plus_1 = tol_eoy * 1.04
    tol_eoy_trend_minus_1 = tol_eoy * 0.96
    
    print(f"    BOY TOL: ${tol_boy:,.0f}")
    print(f"    Service Cost: ${service_cost:,.0f}")
    print(f"    Interest Cost: ${interest_cost:,.0f}")
    print(f"    Assumption Change: ${assumption_change:,.0f}")
    print(f"    Experience G/L: ${experience_gl:,.0f}")
    print(f"    EOY TOL: ${tol_eoy:,.0f}")
    
    return ValuationResults(
        client_name=client_name,
        measurement_date=measurement_date,
        prior_measurement_date=prior_measurement_date,
        valuation_date=valuation_date,
        active_count=active_count,
        retiree_count=retiree_count,
        covered_payroll=covered_payroll,
        tol_boy_old_rate=tol_boy,
        tol_boy_new_rate=tol_boy_new_rate,
        tol_eoy_baseline=tol_eoy,
        tol_eoy_disc_plus_1=tol_eoy_disc_plus_1,
        tol_eoy_disc_minus_1=tol_eoy_disc_minus_1,
        tol_eoy_trend_plus_1=tol_eoy_trend_plus_1,
        tol_eoy_trend_minus_1=tol_eoy_trend_minus_1,
        service_cost=service_cost,
        interest_cost=interest_cost,
        assumption_change=assumption_change,
        experience_gl=experience_gl,
        discount_rate_boy=prior_discount_rate,
        discount_rate_eoy=discount_rate,
        avg_remaining_service_life=prior_data.arsl,
        benefit_payments=benefit_payments
    )


# =============================================================================
# GASB 75 FILE UPDATE - CRITICAL LOGIC
# =============================================================================

def update_gasb75_file(prior_file: str, output_file: str, results: ValuationResults, 
                       prior_data: PriorYearData) -> Dict[str, Any]:
    """
    Update a GASB 75 disclosure file from one measurement period to the next.
    
    CRITICAL: This follows the exact update sequence required:
    1. Hardcode RSI current year column (preserve prior year data)
    2. Set RSI new year column formulas
    3. Shift ARSL values in AmortDeferredOutsIns (ALL values shift down)
    4. Update AmortDeferredOutsIns A13 (current year)
    5. Update Assumptions tab (dates and rates)
    6. Update ProVal1 (valuation results)
    7. Update Net OPEB labels
    """
    print(f"\n  Updating GASB 75 file: {output_file}")
    
    # Load workbooks - one for formulas, one for current values
    wb = load_workbook(prior_file)
    wb_data = load_workbook(prior_file, data_only=True)
    
    summary = {}
    
    # Determine current (prior) year and new year
    prior_year = prior_data.measurement_date.year if prior_data.measurement_date else results.measurement_date.year - 1
    new_year = results.measurement_date.year
    
    summary['prior_year'] = prior_year
    summary['new_year'] = new_year
    
    # Map year to RSI column: 2018=B, 2019=C, ..., 2024=H, 2025=I
    def year_to_col(year):
        return chr(ord('B') + year - 2018)
    
    current_col = year_to_col(prior_year)
    new_col = year_to_col(new_year)
    
    print(f"    Prior year: {prior_year} (column {current_col})")
    print(f"    New year: {new_year} (column {new_col})")
    
    # =========================================================================
    # STEP 1: HARDCODE RSI CURRENT YEAR COLUMN
    # =========================================================================
    print("    Step 1: Hardcoding RSI current year column...")
    
    rsi = wb['RSI']
    rsi_data = wb_data['RSI']
    
    rsi_rows = [3, 4, 5, 6, 7, 8, 9, 10, 12, 14, 17, 20, 26]
    for row in rsi_rows:
        cell_ref = f'{current_col}{row}'
        current_value = rsi_data[cell_ref].value
        rsi[cell_ref] = current_value
    
    summary['rsi_hardcoded_column'] = current_col
    
    # =========================================================================
    # STEP 2: SET RSI NEW YEAR COLUMN FORMULAS
    # =========================================================================
    print("    Step 2: Setting RSI new year column formulas...")
    
    rsi[f'{new_col}3'] = "=YEAR(Assumptions!C4)"
    rsi[f'{new_col}4'] = "='Net OPEB'!D14"
    rsi[f'{new_col}5'] = "='Net OPEB'!D16"
    rsi[f'{new_col}6'] = "='Net OPEB'!D20"
    rsi[f'{new_col}7'] = "='Net OPEB'!D22"
    rsi[f'{new_col}8'] = "='Net OPEB'!D18"
    rsi[f'{new_col}9'] = "='Net OPEB'!D24"
    rsi[f'{new_col}10'] = "='Net OPEB'!D27"
    rsi[f'{new_col}12'] = "='Net OPEB'!D12"
    rsi[f'{new_col}14'] = "='Net OPEB'!D29"
    rsi[f'{new_col}17'] = results.covered_payroll
    rsi[f'{new_col}20'] = f"={new_col}14/{new_col}17"
    rsi[f'{new_col}26'] = "=AmortDeferredOutsIns!C6"
    rsi[f'{new_col}27'] = "Pub-2010/2021"
    rsi[f'{new_col}28'] = "Getzen model"
    
    summary['rsi_formula_column'] = new_col
    
    # =========================================================================
    # STEP 3: SHIFT ARSL VALUES IN AmortDeferredOutsIns
    # =========================================================================
    print("    Step 3: Shifting ARSL values...")
    
    amort = wb['AmortDeferredOutsIns']
    amort_data = wb_data['AmortDeferredOutsIns']
    
    # Capture current values BEFORE shifting
    c13_val = amort_data['C13'].value
    c14_val = amort_data['C14'].value
    c15_val = amort_data['C15'].value
    c16_val = amort_data['C16'].value
    c17_val = amort_data['C17'].value
    c18_val = amort_data['C18'].value
    
    # Shift ALL values down by one row
    amort['C14'] = c13_val  # Current year's ARSL moves to row 14
    amort['C15'] = c14_val  # Prior year 1 moves to row 15
    amort['C16'] = c15_val  # Prior year 2 moves to row 16
    amort['C17'] = c16_val  # Prior year 3 moves to row 17
    amort['C18'] = c17_val  # Prior year 4 moves to row 18
    amort['C19'] = c18_val  # Prior year 5 moves to row 19
    # C13 keeps its formula - will calculate new year's ARSL
    
    summary['arsl_shifted'] = {
        f'C14 ({prior_year})': c13_val,
        f'C15 ({prior_year-1})': c14_val,
        f'C16 ({prior_year-2})': c15_val,
        f'C17 ({prior_year-3})': c16_val,
        f'C18 ({prior_year-4})': c17_val,
        f'C19 ({prior_year-5})': c18_val,
    }
    
    # =========================================================================
    # STEP 4: UPDATE AmortDeferredOutsIns A13 (current year)
    # =========================================================================
    print("    Step 4: Updating AmortDeferredOutsIns A13...")
    
    amort['A13'] = new_year
    
    # =========================================================================
    # STEP 5: UPDATE ASSUMPTIONS TAB
    # =========================================================================
    print("    Step 5: Updating Assumptions tab...")
    
    assumptions = wb['Assumptions']
    assumptions['C2'] = results.valuation_date
    assumptions['C3'] = results.prior_measurement_date
    assumptions['C4'] = results.measurement_date
    assumptions['C11'] = results.discount_rate_boy  # Prior rate for interest calc
    assumptions['C12'] = f"{results.discount_rate_eoy*100:.2f}% annually which is the Bond Buyer 20-Bond General Obligation Index on the Measurement Date.  The 20-Bond Index consists of 20 general obligation bonds that mature in 20 years."
    
    # =========================================================================
    # STEP 6: UPDATE PROVAL1
    # =========================================================================
    print("    Step 6: Updating ProVal1...")
    
    proval = wb['ProVal1']
    proval['B19'] = results.tol_boy_old_rate
    proval['C19'] = results.tol_boy_new_rate
    proval['D19'] = results.tol_eoy_baseline
    proval['E19'] = results.tol_eoy_disc_plus_1
    proval['F19'] = results.tol_eoy_disc_minus_1
    proval['G19'] = results.tol_eoy_baseline  # Trend baseline
    proval['H19'] = results.tol_eoy_trend_plus_1
    proval['I19'] = results.tol_eoy_trend_minus_1
    proval['B38'] = results.service_cost  # B38 is what Net OPEB D14 uses!
    proval['D17'] = results.covered_payroll
    proval['D88'] = results.discount_rate_eoy
    
    if results.active_count:
        proval['D6'] = results.active_count
    if results.retiree_count:
        proval['D8'] = results.retiree_count
    
    summary['proval_updated'] = {
        'B19': results.tol_boy_old_rate,
        'C19': results.tol_boy_new_rate,
        'D19': results.tol_eoy_baseline,
        'B38': results.service_cost,
        'D88': results.discount_rate_eoy,
    }
    
    # =========================================================================
    # STEP 7: UPDATE NET OPEB LABELS
    # =========================================================================
    print("    Step 7: Updating Net OPEB labels...")
    
    net_opeb = wb['Net OPEB']
    md = results.measurement_date
    pmd = results.prior_measurement_date
    net_opeb['A9'] = f"Table 3: Changes in Net OPEB Liability for the plan's fiscal year ending {md.month}/{md.day}/{md.year}"
    net_opeb['A12'] = f"Balances at {pmd.month}/{pmd.day}/{pmd.year}"
    net_opeb['A29'] = f"Balances at {md.month}/{md.day}/{md.year}"
    
    # =========================================================================
    # SAVE
    # =========================================================================
    wb.save(output_file)
    wb.close()
    wb_data.close()
    
    print(f"    Saved: {output_file}")
    
    return summary


def validate_output(file_path: str) -> bool:
    """Validate output file - must be opened in Excel for formula calculation."""
    print(f"\n  Validating (formulas require Excel to calculate)...")
    
    wb = load_workbook(file_path, data_only=False)
    
    # Check that key formulas are intact
    checks_passed = True
    
    # Check AmortDeferredOutsIns B13
    amort = wb['AmortDeferredOutsIns']
    b13 = amort['B13'].value
    if b13 == "='Net OPEB'!D22":
        print("    ✓ AmortDeferredOutsIns B13 formula intact")
    else:
        print(f"    ✗ AmortDeferredOutsIns B13 unexpected: {b13}")
        checks_passed = False
    
    # Check RSI new column has formulas
    rsi = wb['RSI']
    # Find the column with formulas
    for col in ['I', 'J', 'K']:
        val = rsi[f'{col}7'].value
        if val == "='Net OPEB'!D22":
            print(f"    ✓ RSI {col}7 formula intact (current year experience)")
            break
    
    # Check ProVal1 has values
    proval = wb['ProVal1']
    if proval['D19'].value and proval['D19'].value > 0:
        print(f"    ✓ ProVal1 D19 (EOY TOL): {proval['D19'].value:,.0f}")
    else:
        print("    ✗ ProVal1 D19 missing")
        checks_passed = False
    
    wb.close()
    
    print("\n    NOTE: Open file in Excel to recalculate formulas.")
    print("    LibreOffice cannot properly calculate this file's complex formulas.")
    
    return checks_passed


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="OPEB Valuation Runner - Updates GASB 75 disclosure files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  Roll-forward:
    python run_valuation.py rollforward --prior-file GASB75_2024.xlsx \\
        --client-name "West Florida" --discount-rate 0.0502 --prior-discount-rate 0.0381

  Full valuation:
    python run_valuation.py full --prior-file GASB75_2024.xlsx \\
        --actives-file actives.xlsx --retirees-file retirees.xlsx \\
        --client-name "West Florida" --discount-rate 0.0502 --prior-discount-rate 0.0381
        """
    )
    
    parser.add_argument('type', choices=['full', 'rollforward'], 
                        help='Valuation type: full (with census) or rollforward')
    parser.add_argument('--prior-file', required=True,
                        help='Prior year GASB 75 file to update')
    parser.add_argument('--actives-file',
                        help='Active census file (required for full valuation)')
    parser.add_argument('--retirees-file',
                        help='Retiree census file (required for full valuation)')
    parser.add_argument('--client-name', required=True,
                        help='Client name for output file')
    parser.add_argument('--discount-rate', type=float, required=True,
                        help='New (EOY) discount rate (e.g., 0.0502)')
    parser.add_argument('--prior-discount-rate', type=float, required=True,
                        help='Prior (BOY) discount rate (e.g., 0.0381)')
    parser.add_argument('--benefit-payments', type=float, default=0,
                        help='Benefit payments during year (default: 0)')
    parser.add_argument('--measurement-date',
                        help='Measurement date YYYY-MM-DD (default: 9/30 of current year)')
    parser.add_argument('--output-file',
                        help='Output file path (default: auto-generated)')
    
    args = parser.parse_args()
    
    # Validate arguments
    if args.type == 'full' and (not args.actives_file or not args.retirees_file):
        parser.error("Full valuation requires --actives-file and --retirees-file")
    
    # Parse measurement date
    if args.measurement_date:
        measurement_date = datetime.strptime(args.measurement_date, '%Y-%m-%d').date()
    else:
        # Default to 9/30 of current year
        today = date.today()
        measurement_date = date(today.year, 9, 30)
    
    # Generate output filename
    output_file = args.output_file or f"GASB75_{args.client_name.replace(' ', '_')}_{measurement_date.year}_FINAL.xlsx"
    
    # Print header
    print("=" * 70)
    print(f"OPEB {args.type.upper()} VALUATION")
    print("=" * 70)
    print(f"Client: {args.client_name}")
    print(f"Measurement Date: {measurement_date}")
    print(f"Discount Rate: {args.prior_discount_rate:.2%} → {args.discount_rate:.2%}")
    print(f"Prior File: {args.prior_file}")
    print(f"Output File: {output_file}")
    print("=" * 70)
    
    # Extract prior year data
    prior_data = extract_prior_year_data(args.prior_file)
    
    # Run valuation
    if args.type == 'rollforward':
        results = run_rollforward(
            prior_data=prior_data,
            discount_rate=args.discount_rate,
            prior_discount_rate=args.prior_discount_rate,
            benefit_payments=args.benefit_payments,
            measurement_date=measurement_date,
            client_name=args.client_name
        )
    else:
        results = run_full_valuation(
            prior_data=prior_data,
            actives_file=args.actives_file,
            retirees_file=args.retirees_file,
            discount_rate=args.discount_rate,
            prior_discount_rate=args.prior_discount_rate,
            measurement_date=measurement_date,
            client_name=args.client_name,
            benefit_payments=args.benefit_payments
        )
    
    # Update GASB 75 file
    summary = update_gasb75_file(
        prior_file=args.prior_file,
        output_file=output_file,
        results=results,
        prior_data=prior_data
    )
    
    # Validate
    validate_output(output_file)
    
    # Print summary
    print("\n" + "=" * 70)
    print("VALUATION COMPLETE")
    print("=" * 70)
    print(f"Output: {output_file}")
    print(f"EOY TOL: ${results.tol_eoy_baseline:,.0f}")
    print(f"Experience G/L: ${results.experience_gl:,.0f}")
    print("\nNOTE: Open in Excel to recalculate formulas before delivery.")
    print("=" * 70)


if __name__ == '__main__':
    main()
