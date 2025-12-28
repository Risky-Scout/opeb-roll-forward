"""
opeb_rollforward/excel_updater.py - Production GASB 75 Excel Template Updater

Version: 2025-12-28 (West Florida Planning Corrections)

This module handles updating existing GASB 75 Excel disclosure templates
for roll-forward valuations. It incorporates all corrections developed
during production debugging.

CRITICAL FIXES INCLUDED:
1. Clear OPEB Exp & Def C6:C28 before starting
2. Net OPEB D22:D25 must have No Fill
3. RSI I23 = "None" for roll-forwards (or benefit change description)
4. RSI I26 = discount rate VALUE (not formula reference)
5. Table7AmortDeferred B14/B26 must copy FULL cell style from adjacent cells
6. Table7AmortDeferred B13 formula forces near-zero experience to $0
7. Skip Table7AmortDeferred2 (doesn't serve purpose currently)

Author: Actuarial Pipeline Project
License: MIT
"""

from openpyxl import load_workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import date
from dataclasses import dataclass
from typing import Optional, Dict, Any, Tuple
import re
import logging

logger = logging.getLogger(__name__)


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def copy_cell_format(source_cell, target_cell):
    """
    Copy COMPLETE cell formatting from source to target.
    
    Copies: font, border, fill, number_format, protection, alignment.
    
    CRITICAL: Use this instead of just setting number_format to ensure
    cells match surrounding formatting exactly.
    """
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def adjust_formula_row(formula, old_row: int, new_row: int) -> str:
    """
    Adjust formula row references from old_row to new_row.
    
    Example:
        adjust_formula_row("=IF(D$11>=$A14,$B14/$C14,0)", 14, 13)
        Returns: "=IF(D$11>=$A13,$B13/$C13,0)"
    """
    if formula is None:
        return None
    if not isinstance(formula, str):
        return formula
    pattern = r'(\$?[A-Z]+)' + str(old_row) + r'(?![0-9])'
    replacement = r'\g<1>' + str(new_row)
    return re.sub(pattern, replacement, formula)


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class RollForwardInputs:
    """Input parameters for a roll-forward valuation."""
    # Dates
    prior_measurement_date: date
    new_measurement_date: date
    
    # Discount rates
    prior_discount_rate: float
    new_discount_rate: float
    
    # Liabilities and costs (from prior year)
    boy_tol_old_rate: float  # BOY liability at prior year's EOY rate
    service_cost: float       # Prior year service cost (used for roll-forward)
    
    # Payroll
    covered_payroll_prior: float
    
    # Durations (for sensitivity calculations)
    duration: float = 10.0
    trend_duration: float = 5.0
    
    # Growth rates
    payroll_growth_rate: float = 0.03
    
    # Benefit changes description (or "None")
    benefit_changes: str = "None"


@dataclass
class RollForwardResults:
    """Calculated results from a roll-forward valuation."""
    # Liabilities
    boy_tol_old_rate: float
    boy_tol_new_rate: float
    eoy_tol: float
    
    # Components
    service_cost: float
    interest: float
    assumption_change: float
    experience: float  # Always 0 for pure roll-forward
    
    # Sensitivities
    sensitivity_disc_plus: float
    sensitivity_disc_minus: float
    sensitivity_trend_plus: float
    sensitivity_trend_minus: float
    
    # Payroll
    covered_payroll_new: float


# =============================================================================
# ACTUARIAL CALCULATIONS
# =============================================================================

def calculate_roll_forward(inputs: RollForwardInputs) -> RollForwardResults:
    """
    Perform roll-forward actuarial calculations.
    
    Uses duration approximation for assumption changes and mid-year
    discounting for interest calculations.
    
    For roll-forwards:
    - Experience = $0 (no census data)
    - Assumption change = effect of discount rate change
    """
    # Interest = (BOY + 0.5 * SC) * prior_rate (mid-year approximation)
    interest = (inputs.boy_tol_old_rate + 0.5 * inputs.service_cost) * inputs.prior_discount_rate
    
    # BOY at new rate using duration approximation
    rate_change = inputs.new_discount_rate - inputs.prior_discount_rate
    boy_tol_new_rate = inputs.boy_tol_old_rate * (1 - inputs.duration * rate_change)
    assumption_change = boy_tol_new_rate - inputs.boy_tol_old_rate
    
    # EOY = BOY + SC + Interest + Assumption (Experience = 0 for roll-forward)
    experience = 0.0
    eoy_tol = inputs.boy_tol_old_rate + inputs.service_cost + interest + assumption_change
    
    # Sensitivities using duration approximation
    sensitivity_disc_plus = eoy_tol * (1 - inputs.duration * 0.01)
    sensitivity_disc_minus = eoy_tol * (1 - inputs.duration * -0.01)
    sensitivity_trend_plus = eoy_tol * (1 + inputs.trend_duration * 0.01)
    sensitivity_trend_minus = eoy_tol * (1 + inputs.trend_duration * -0.01)
    
    # Covered payroll growth
    covered_payroll_new = inputs.covered_payroll_prior * (1 + inputs.payroll_growth_rate)
    
    return RollForwardResults(
        boy_tol_old_rate=inputs.boy_tol_old_rate,
        boy_tol_new_rate=boy_tol_new_rate,
        eoy_tol=eoy_tol,
        service_cost=inputs.service_cost,
        interest=interest,
        assumption_change=assumption_change,
        experience=experience,
        sensitivity_disc_plus=sensitivity_disc_plus,
        sensitivity_disc_minus=sensitivity_disc_minus,
        sensitivity_trend_plus=sensitivity_trend_plus,
        sensitivity_trend_minus=sensitivity_trend_minus,
        covered_payroll_new=covered_payroll_new,
    )


# =============================================================================
# EXCEL UPDATE FUNCTIONS
# =============================================================================

def update_roll_forward_excel(
    input_path: str,
    output_path: str,
    inputs: RollForwardInputs,
    results: Optional[RollForwardResults] = None,
) -> str:
    """
    Update GASB 75 Excel template with roll-forward results.
    
    Implements all corrections from the West Florida Planning debugging session.
    
    Args:
        input_path: Path to prior year Excel file
        output_path: Path for output file
        inputs: RollForwardInputs
        results: RollForwardResults (calculated if not provided)
        
    Returns:
        Path to saved output file
    """
    # Calculate results if not provided
    if results is None:
        results = calculate_roll_forward(inputs)
    
    # Load workbook twice - once for formulas, once for values
    wb = load_workbook(input_path)
    wb_data = load_workbook(input_path, data_only=True)
    
    measurement_year = inputs.new_measurement_date.year
    
    logger.info(f"Starting roll-forward update: {inputs.prior_measurement_date} → {inputs.new_measurement_date}")
    
    # =========================================================================
    # STEP 0: Remove all conditional formatting
    # =========================================================================
    logger.info("Step 0: Removing conditional formatting")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.conditional_formatting = ConditionalFormattingList()
    
    # =========================================================================
    # STEP 0.5: Clear OPEB Exp & Def C6:C28 (CRITICAL!)
    # =========================================================================
    logger.info("Step 0.5: Clearing OPEB Exp & Def C6:C28")
    opeb_exp = wb['OPEB Exp & Def']
    for row in range(6, 29):
        opeb_exp[f'C{row}'] = None
    
    # =========================================================================
    # STEP 1: Update Model Inputs Tab
    # =========================================================================
    logger.info("Step 1: Updating Model Inputs tab")
    mi = wb['Model Inputs']
    
    # TOL values
    mi['B19'] = results.boy_tol_old_rate
    mi['C19'] = round(results.boy_tol_new_rate)
    mi['D19'] = round(results.eoy_tol)
    mi['E19'] = round(results.sensitivity_disc_plus)
    mi['F19'] = round(results.sensitivity_disc_minus)
    mi['G19'] = round(results.eoy_tol)
    mi['H19'] = round(results.sensitivity_trend_plus)
    mi['I19'] = round(results.sensitivity_trend_minus)
    
    # Service cost
    mi['B38'] = results.service_cost
    mi['C38'] = results.service_cost
    mi['D38'] = results.service_cost
    
    # Covered payroll
    mi['B17'] = inputs.covered_payroll_prior
    mi['C17'] = inputs.covered_payroll_prior
    for col in 'DEFGHI':
        mi[f'{col}17'] = round(results.covered_payroll_new)
    
    # Dates
    mi['B77'] = inputs.prior_measurement_date
    mi['C77'] = inputs.prior_measurement_date
    for col in 'DEFGHI':
        mi[f'{col}77'] = inputs.new_measurement_date
    
    # Discount rates
    mi['B88'] = inputs.prior_discount_rate
    mi['C88'] = inputs.new_discount_rate
    mi['D88'] = inputs.new_discount_rate
    
    # =========================================================================
    # STEP 2: Set Net OPEB D22:D25 to No Fill (CRITICAL!)
    # =========================================================================
    logger.info("Step 2: Setting Net OPEB D22:D25 to no fill")
    net_opeb = wb['Net OPEB']
    no_fill = PatternFill(fill_type=None)
    for row in range(22, 26):
        net_opeb[f'D{row}'].fill = no_fill
    
    # =========================================================================
    # STEP 3: Update RSI Tab (WITH FORMATTING)
    # =========================================================================
    logger.info("Step 3: Updating RSI tab")
    rsi = wb['RSI']
    rsi_data = wb_data['RSI']
    
    # Include row 23 in the update rows!
    rsi_rows = [3, 4, 5, 6, 7, 8, 9, 10, 12, 14, 17, 20, 23, 26, 27, 28]
    
    # Get current H column values
    h_values = {}
    for row in rsi_rows:
        h_values[row] = rsi_data[f'H{row}'].value
    
    # Copy formatting from H to I for ALL rows
    for row in rsi_rows:
        h_cell = rsi[f'H{row}']
        i_cell = rsi[f'I{row}']
        copy_cell_format(h_cell, i_cell)
    
    # Set I column values/formulas
    rsi['I3'] = "=YEAR(Assumptions!C4)"
    rsi['I4'] = "='Net OPEB'!D14"
    rsi['I5'] = "='Net OPEB'!D16"
    rsi['I6'] = "='Net OPEB'!D20"
    rsi['I7'] = "='Net OPEB'!D22"
    rsi['I8'] = "='Net OPEB'!D18"
    rsi['I9'] = "='Net OPEB'!D24"
    rsi['I10'] = "='Net OPEB'!D27"
    rsi['I12'] = "='Net OPEB'!D12"
    rsi['I14'] = "='Net OPEB'!D29"
    rsi['I17'] = round(results.covered_payroll_new)
    rsi['I20'] = "=I14/I17"
    
    # CRITICAL: I23 = "None" for roll-forwards
    rsi['I23'] = inputs.benefit_changes
    
    # CRITICAL: I26 = discount rate as VALUE, NOT a formula!
    rsi['I26'] = inputs.new_discount_rate
    
    rsi['I27'] = "Pub-2010/2021"
    rsi['I28'] = "Getzen model"
    
    # Hardcode H column with prior year values
    for row in rsi_rows:
        rsi[f'H{row}'] = h_values[row]
    
    # =========================================================================
    # STEP 4: Update Table7AmortDeferred (WITH FULL FORMATTING)
    # =========================================================================
    logger.info("Step 4: Updating Table7AmortDeferred")
    t7 = wb['Table7AmortDeferred']
    t7_data = wb_data['Table7AmortDeferred']
    
    # SECTION 1: Experience (row 14 → row 13)
    logger.info("  Section 1: Experience amortization")
    
    copy_cell_format(t7['A14'], t7['A13'])
    t7['A13'] = measurement_year
    
    copy_cell_format(t7['B14'], t7['B13'])
    # CRITICAL: Formula forces near-zero experience to exactly $0
    t7['B13'] = "=IF('Net OPEB'!D22<1,0,'Net OPEB'!D22)"
    
    copy_cell_format(t7['C14'], t7['C13'])
    t7['C13'] = t7['C14'].value
    
    # CRITICAL: Hardcode B14, then copy FULL style from B15
    b14_value = t7_data['B14'].value
    t7['B14'] = round(b14_value) if b14_value else 0
    copy_cell_format(t7['B15'], t7['B14'])  # Copy from adjacent cell!
    
    c14_value = t7_data['C14'].value
    t7['C14'] = round(c14_value) if c14_value else 0
    
    # Copy D14:AI14 → D13:AI13 with row adjustment
    for col_idx in range(4, 36):
        col = get_column_letter(col_idx)
        copy_cell_format(t7[f'{col}14'], t7[f'{col}13'])
        formula = t7[f'{col}14'].value
        if formula and isinstance(formula, str) and formula.startswith('='):
            t7[f'{col}13'].value = adjust_formula_row(formula, 14, 13)
    
    t7['AJ14'] = 1
    
    # SECTION 2: Assumptions (row 26 → row 25)
    logger.info("  Section 2: Assumptions amortization")
    
    copy_cell_format(t7['A26'], t7['A25'])
    t7['A25'] = measurement_year
    
    copy_cell_format(t7['B26'], t7['B25'])
    t7['B25'] = t7['B26'].value
    
    copy_cell_format(t7['C26'], t7['C25'])
    c26_formula = t7['C26'].value
    if c26_formula and isinstance(c26_formula, str):
        t7['C25'] = adjust_formula_row(c26_formula, 14, 13)
    
    # CRITICAL: Hardcode B26, then copy FULL style from B27
    b26_value = t7_data['B26'].value
    t7['B26'] = round(b26_value) if b26_value else 0
    copy_cell_format(t7['B27'], t7['B26'])  # Copy from adjacent cell!
    
    c26_value = t7_data['C26'].value
    t7['C26'] = round(c26_value) if c26_value else 0
    
    # Copy D26:AI26 → D25:AI25 with row adjustment
    for col_idx in range(4, 36):
        col = get_column_letter(col_idx)
        copy_cell_format(t7[f'{col}26'], t7[f'{col}25'])
        formula = t7[f'{col}26'].value
        if formula and isinstance(formula, str) and formula.startswith('='):
            t7[f'{col}25'].value = adjust_formula_row(formula, 26, 25)
    
    t7['AJ26'] = 1
    
    # SECTION 3: Total (row 38 → row 37)
    logger.info("  Section 3: Total amortization")
    
    copy_cell_format(t7['A38'], t7['A37'])
    t7['A37'] = measurement_year
    
    copy_cell_format(t7['B38'], t7['B37'])
    t7['B37'] = "=B13+B25"
    
    # Copy D38:AI38 → D37:AI37 with MULTIPLE row adjustments
    for col_idx in range(4, 36):
        col = get_column_letter(col_idx)
        copy_cell_format(t7[f'{col}38'], t7[f'{col}37'])
        formula = t7[f'{col}38'].value
        if formula and isinstance(formula, str) and formula.startswith('='):
            adjusted = adjust_formula_row(formula, 38, 37)
            adjusted = adjust_formula_row(adjusted, 14, 13)
            adjusted = adjust_formula_row(adjusted, 26, 25)
            t7[f'{col}37'].value = adjusted
    
    t7['AJ38'] = 1
    
    # =========================================================================
    # STEP 5: SKIP Table7AmortDeferred2 (doesn't serve purpose)
    # =========================================================================
    logger.info("Step 5: Skipping Table7AmortDeferred2")
    
    # =========================================================================
    # STEP 6: Update AmortDeferredOutsIns Tab
    # =========================================================================
    logger.info("Step 6: Updating AmortDeferredOutsIns")
    amort = wb['AmortDeferredOutsIns']
    amort_data = wb_data['AmortDeferredOutsIns']
    
    amort['A13'] = measurement_year
    
    # Shift Experience ARSL (C13→C14, C14→C15, etc.)
    for src, dst in [(13, 14), (14, 15), (15, 16), (16, 17), (17, 18), (18, 19)]:
        amort[f'C{dst}'] = amort_data[f'C{src}'].value
    
    # Shift Assumptions ARSL (C23→C24, C24→C25, etc.)
    for src, dst in [(23, 24), (24, 25), (25, 26), (26, 27), (27, 28), (28, 29)]:
        amort[f'C{dst}'] = amort_data[f'C{src}'].value
    
    # =========================================================================
    # SAVE
    # =========================================================================
    logger.info(f"Saving to: {output_path}")
    wb.save(output_path)
    wb.close()
    wb_data.close()
    
    logger.info("Roll-forward complete!")
    
    return output_path


# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

def run_roll_forward(
    input_path: str,
    output_path: str,
    prior_measurement_date: date,
    new_measurement_date: date,
    prior_discount_rate: float,
    new_discount_rate: float,
    duration: float = 10.0,
    trend_duration: float = 5.0,
    payroll_growth_rate: float = 0.03,
    benefit_changes: str = "None",
) -> Tuple[str, RollForwardResults]:
    """
    Run a complete roll-forward valuation.
    
    Extracts prior year data from input file, calculates roll-forward,
    and updates the Excel template.
    
    Args:
        input_path: Path to prior year GASB 75 Excel file
        output_path: Path for output file
        prior_measurement_date: Prior measurement date (e.g., 9/30/2024)
        new_measurement_date: New measurement date (e.g., 9/30/2025)
        prior_discount_rate: Prior year EOY discount rate (e.g., 0.0381)
        new_discount_rate: New year discount rate (e.g., 0.0502)
        duration: Liability duration for sensitivity (default 10.0)
        trend_duration: Healthcare trend duration (default 5.0)
        payroll_growth_rate: Annual payroll growth (default 0.03)
        benefit_changes: Description of benefit changes or "None"
        
    Returns:
        Tuple of (output_path, RollForwardResults)
    """
    # Load prior year data
    wb_data = load_workbook(input_path, data_only=True)
    mi = wb_data['Model Inputs']
    
    boy_tol_old_rate = mi['D19'].value
    service_cost = mi['D38'].value
    covered_payroll_prior = mi['D17'].value
    
    wb_data.close()
    
    # Create inputs
    inputs = RollForwardInputs(
        prior_measurement_date=prior_measurement_date,
        new_measurement_date=new_measurement_date,
        prior_discount_rate=prior_discount_rate,
        new_discount_rate=new_discount_rate,
        boy_tol_old_rate=boy_tol_old_rate,
        service_cost=service_cost,
        covered_payroll_prior=covered_payroll_prior,
        duration=duration,
        trend_duration=trend_duration,
        payroll_growth_rate=payroll_growth_rate,
        benefit_changes=benefit_changes,
    )
    
    # Calculate results
    results = calculate_roll_forward(inputs)
    
    # Update Excel
    output = update_roll_forward_excel(input_path, output_path, inputs, results)
    
    return output, results


def print_roll_forward_summary(results: RollForwardResults, inputs: RollForwardInputs):
    """Print a formatted summary of roll-forward results."""
    print("=" * 60)
    print("GASB 75 ROLL-FORWARD SUMMARY")
    print("=" * 60)
    print(f"Measurement Period: {inputs.prior_measurement_date} → {inputs.new_measurement_date}")
    print(f"Discount Rate: {inputs.prior_discount_rate:.2%} → {inputs.new_discount_rate:.2%}")
    print()
    print("TOL Reconciliation:")
    print(f"  BOY TOL (old rate):     ${results.boy_tol_old_rate:>12,.0f}")
    print(f"  BOY TOL (new rate):     ${results.boy_tol_new_rate:>12,.0f}")
    print(f"  Service Cost:           ${results.service_cost:>12,.0f}")
    print(f"  Interest:               ${results.interest:>12,.0f}")
    print(f"  Assumption Change:      ${results.assumption_change:>12,.0f}")
    print(f"  Experience:             ${results.experience:>12,.0f}")
    print(f"  EOY TOL:                ${results.eoy_tol:>12,.0f}")
    print()
    print("Sensitivities:")
    print(f"  Discount +1%:           ${results.sensitivity_disc_plus:>12,.0f}")
    print(f"  Discount -1%:           ${results.sensitivity_disc_minus:>12,.0f}")
    print(f"  Trend +1%:              ${results.sensitivity_trend_plus:>12,.0f}")
    print(f"  Trend -1%:              ${results.sensitivity_trend_minus:>12,.0f}")
    print()
    print(f"Covered Payroll:          ${results.covered_payroll_new:>12,.0f}")
    print("=" * 60)


# =============================================================================
# VERIFICATION FUNCTIONS
# =============================================================================

def verify_roll_forward_output(output_path: str) -> Dict[str, Any]:
    """
    Verify the roll-forward output meets all quality checks.
    
    Returns dict with verification results.
    """
    wb = load_workbook(output_path, data_only=True)
    
    results = {
        'passed': True,
        'checks': {}
    }
    
    # Check 1: Table7AmortDeferred AI49 should be "GOOD"
    t7 = wb['Table7AmortDeferred']
    ai49 = t7['AI49'].value
    results['checks']['Table7AmortDeferred_AI49'] = {
        'expected': 'GOOD',
        'actual': ai49,
        'passed': str(ai49).upper() == 'GOOD' if ai49 else False
    }
    
    # Check 2: OPEB Exp & Def H40 should be "GOOD"
    opeb = wb['OPEB Exp & Def']
    h40 = opeb['H40'].value
    results['checks']['OPEB_Exp_Def_H40'] = {
        'expected': 'GOOD',
        'actual': h40,
        'passed': str(h40).upper() == 'GOOD' if h40 else False
    }
    
    # Check 3: RSI I26 should be a number (discount rate)
    rsi = wb['RSI']
    i26 = rsi['I26'].value
    results['checks']['RSI_I26_is_number'] = {
        'expected': 'numeric value',
        'actual': i26,
        'passed': isinstance(i26, (int, float))
    }
    
    # Check 4: RSI I23 should have a value
    i23 = rsi['I23'].value
    results['checks']['RSI_I23_has_value'] = {
        'expected': 'non-empty',
        'actual': i23,
        'passed': i23 is not None and str(i23).strip() != ''
    }
    
    wb.close()
    
    # Overall pass/fail
    results['passed'] = all(c['passed'] for c in results['checks'].values())
    
    return results


# =============================================================================
# MAIN
# =============================================================================

if __name__ == '__main__':
    import sys
    
    print("GASB 75 Roll-Forward Excel Updater")
    print("=" * 50)
    print()
    print("Usage:")
    print("  from opeb_rollforward.excel_updater import run_roll_forward")
    print()
    print("  output, results = run_roll_forward(")
    print("      input_path='prior_year.xlsx',")
    print("      output_path='current_year.xlsx',")
    print("      prior_measurement_date=date(2024, 9, 30),")
    print("      new_measurement_date=date(2025, 9, 30),")
    print("      prior_discount_rate=0.0381,")
    print("      new_discount_rate=0.0502,")
    print("  )")
