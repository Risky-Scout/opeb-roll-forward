#!/usr/bin/env python3
"""
GASB 75 OPEB Disclosure Excel File Updater
==========================================

This module contains the correct logic for updating a GASB 75 disclosure file
from one measurement period to the next.

CRITICAL UPDATE SEQUENCE:
1. Hardcode RSI current year column (preserve prior year data)
2. Set RSI new year column formulas  
3. Shift ARSL values in AmortDeferredOutsIns (ALL values shift down)
4. Update AmortDeferredOutsIns A13 (current year)
5. Update Assumptions tab (dates and rates)
6. Update ProVal1 (valuation results)
7. Update Net OPEB labels

KEY INSIGHTS:
- ARSL values are tied to YEARS, not ROWS. When years shift down, ARSL must follow.
- RSI current year column has formulas to Net OPEB. Must hardcode before updating.
- Interest is calculated at PRIOR discount rate (Assumptions C11).
- Experience is a RESIDUAL in Net OPEB D22.
- Service cost comes from ProVal1 B38 (not D38).

This works for both roll-forward and full valuations - only the ProVal1 values differ.
"""

from openpyxl import load_workbook
from datetime import date
from typing import Dict, Any, Optional, List
from dataclasses import dataclass, field


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class GASB75UpdateInputs:
    """All inputs needed to update a GASB 75 file."""
    # Dates
    valuation_date: date
    prior_measurement_date: date
    measurement_date: date
    
    # Discount rates
    prior_discount_rate: float  # Rate at prior measurement date (for interest calc)
    new_discount_rate: float    # Rate at new measurement date
    
    # ProVal1 values - TOL
    tol_boy_old_rate: float     # B19 - BOY TOL at prior EOY rate
    tol_boy_new_rate: float     # C19 - BOY TOL revalued at new BOY rate
    tol_eoy_baseline: float     # D19 - EOY TOL at new rate
    tol_eoy_disc_plus_1: float  # E19 - EOY at discount +1%
    tol_eoy_disc_minus_1: float # F19 - EOY at discount -1%
    tol_eoy_trend_baseline: float  # G19 - EOY trend baseline
    tol_eoy_trend_plus_1: float    # H19 - EOY at trend +1%
    tol_eoy_trend_minus_1: float   # I19 - EOY at trend -1%
    
    # ProVal1 values - Other
    service_cost: float         # B38 - Service cost for the year
    covered_payroll: float      # D17 - Covered payroll
    active_count: Optional[int] = None   # D6
    retiree_count: Optional[int] = None  # D8


@dataclass 
class PriorYearExtract:
    """Data extracted from prior year GASB 75 file."""
    measurement_date: Optional[date] = None
    tol_eoy: float = 0.0
    service_cost: float = 0.0  # From B38
    discount_rate: float = 0.0
    arsl: float = 5.0
    active_count: int = 0
    retiree_count: int = 0
    covered_payroll: float = 0.0
    # Current ARSL values (will be shifted during update)
    arsl_c13: Optional[float] = None
    arsl_c14: Optional[float] = None
    arsl_c15: Optional[float] = None
    arsl_c16: Optional[float] = None
    arsl_c17: Optional[float] = None
    arsl_c18: Optional[float] = None


# =============================================================================
# EXTRACTION
# =============================================================================

def extract_prior_year_data(prior_file: str) -> PriorYearExtract:
    """
    Extract all required data from prior year GASB 75 file.
    
    Args:
        prior_file: Path to the prior year GASB 75 Excel file
        
    Returns:
        PriorYearExtract with all extracted data
    """
    wb = load_workbook(prior_file, data_only=True)
    data = PriorYearExtract()
    
    # Assumptions - Get measurement date
    assumptions = wb['Assumptions']
    measurement_date = assumptions['C4'].value
    if hasattr(measurement_date, 'date'):
        data.measurement_date = measurement_date.date()
    elif hasattr(measurement_date, 'year'):
        data.measurement_date = measurement_date
    
    # ProVal1 - Key values
    proval = wb['ProVal1']
    data.tol_eoy = proval['D19'].value or 0
    data.service_cost = proval['B38'].value or 0  # B38 is what Net OPEB D14 uses!
    data.active_count = proval['D6'].value or 0
    data.retiree_count = proval['D8'].value or 0
    data.covered_payroll = proval['D17'].value or 0
    data.discount_rate = proval['D88'].value or 0
    data.arsl = proval['D74'].value or 5.0
    
    # AmortDeferredOutsIns - Get ARSL values for shifting
    amort = wb['AmortDeferredOutsIns']
    data.arsl_c13 = amort['C13'].value
    data.arsl_c14 = amort['C14'].value
    data.arsl_c15 = amort['C15'].value
    data.arsl_c16 = amort['C16'].value
    data.arsl_c17 = amort['C17'].value
    data.arsl_c18 = amort['C18'].value
    
    wb.close()
    return data


# =============================================================================
# ROLL-FORWARD CALCULATION
# =============================================================================

def calculate_rollforward(
    prior_tol_eoy: float,
    prior_service_cost: float,
    prior_discount_rate: float,
    new_discount_rate: float,
    benefit_payments: float = 0.0,
    duration: float = 10.0
) -> Dict[str, float]:
    """
    Calculate roll-forward values.
    
    For a roll-forward, we project the liability forward assuming:
    - Service cost same as prior year
    - Interest at prior discount rate (BOY rate)
    - Assumption change from discount rate change
    - Experience = 0 (no actual data)
    
    Args:
        prior_tol_eoy: Prior year's ending TOL (becomes BOY)
        prior_service_cost: Prior year's service cost (from B38)
        prior_discount_rate: Rate at BOY (for interest calculation)
        new_discount_rate: Rate at EOY
        benefit_payments: Benefits paid during year
        duration: Approximate duration for assumption change calc
    
    Returns:
        Dict with all calculated values
    """
    tol_boy = prior_tol_eoy
    service_cost = prior_service_cost
    
    # Interest at prior (BOY) rate
    interest = (tol_boy + 0.5 * service_cost) * prior_discount_rate
    
    # Assumption change from rate change
    rate_change = new_discount_rate - prior_discount_rate
    assumption_change = -tol_boy * rate_change * duration if abs(rate_change) > 0.0001 else 0
    
    # BOY at new rate (for Excel D18 formula: C19 - D12)
    tol_boy_new_rate = tol_boy + assumption_change
    
    # Experience = 0 for roll-forward
    experience = 0.0
    
    # EOY = BOY + SC + Interest + Assumption + Experience - Benefits
    tol_eoy = tol_boy + service_cost + interest + assumption_change + experience - benefit_payments
    
    # Sensitivities (approximate)
    tol_eoy_disc_plus_1 = tol_eoy * 0.92
    tol_eoy_disc_minus_1 = tol_eoy * 1.08
    tol_eoy_trend_plus_1 = tol_eoy * 1.04
    tol_eoy_trend_minus_1 = tol_eoy * 0.96
    
    return {
        'tol_boy_old_rate': tol_boy,
        'tol_boy_new_rate': tol_boy_new_rate,
        'tol_eoy_baseline': tol_eoy,
        'tol_eoy_disc_plus_1': tol_eoy_disc_plus_1,
        'tol_eoy_disc_minus_1': tol_eoy_disc_minus_1,
        'tol_eoy_trend_baseline': tol_eoy,
        'tol_eoy_trend_plus_1': tol_eoy_trend_plus_1,
        'tol_eoy_trend_minus_1': tol_eoy_trend_minus_1,
        'service_cost': service_cost,
        'interest': interest,
        'assumption_change': assumption_change,
        'experience': experience,
    }


# =============================================================================
# FILE UPDATE
# =============================================================================

def update_gasb75_file(
    input_file: str,
    output_file: str,
    inputs: GASB75UpdateInputs,
) -> Dict[str, Any]:
    """
    Update a GASB 75 disclosure file from one measurement period to the next.
    
    CRITICAL: Follows the exact update sequence:
    1. Hardcode RSI current year column
    2. Set RSI new year column formulas
    3. Shift ARSL values (ALL values shift down)
    4. Update A13
    5. Update Assumptions
    6. Update ProVal1
    7. Update Net OPEB labels
    
    Args:
        input_file: Path to the current (prior year) GASB 75 file
        output_file: Path for the updated file
        inputs: GASB75UpdateInputs with all required values
    
    Returns:
        Dict with summary of changes made
    """
    # Load workbooks
    wb = load_workbook(input_file)
    wb_data = load_workbook(input_file, data_only=True)
    
    summary = {'steps_completed': []}
    
    # Determine years
    assumptions_data = wb_data['Assumptions']
    current_measurement_date = assumptions_data['C4'].value
    if hasattr(current_measurement_date, 'year'):
        prior_year = current_measurement_date.year
    else:
        prior_year = inputs.measurement_date.year - 1
    
    new_year = inputs.measurement_date.year
    
    summary['prior_year'] = prior_year
    summary['new_year'] = new_year
    
    # Map year to RSI column
    def year_to_col(year):
        return chr(ord('B') + year - 2018)
    
    current_col = year_to_col(prior_year)
    new_col = year_to_col(new_year)
    
    # =========================================================================
    # STEP 1: HARDCODE RSI CURRENT YEAR COLUMN
    # =========================================================================
    rsi = wb['RSI']
    rsi_data = wb_data['RSI']
    
    rsi_rows = [3, 4, 5, 6, 7, 8, 9, 10, 12, 14, 17, 20, 26]
    hardcoded_values = {}
    for row in rsi_rows:
        cell_ref = f'{current_col}{row}'
        current_value = rsi_data[cell_ref].value
        rsi[cell_ref] = current_value
        hardcoded_values[cell_ref] = current_value
    
    summary['steps_completed'].append('RSI hardcoded')
    summary['rsi_hardcoded'] = hardcoded_values
    
    # =========================================================================
    # STEP 2: SET RSI NEW YEAR COLUMN FORMULAS
    # =========================================================================
    rsi[f'{new_col}3'] = "=YEAR(Assumptions!C4)"
    rsi[f'{new_col}4'] = "='Net OPEB'!D14"
    rsi[f'{new_col}5'] = "='Net OPEB'!D16"
    rsi[f'{new_col}6'] = "='Net OPEB'!D20"
    rsi[f'{new_col}7'] = "='Net OPEB'!D22"
    rsi[f'{new_col}8'] = "='Net OPEB'!D18"
    rsi[f'{new_col}9'] = "='Net OPEB'!D24"
    rsi[f'{new_col}10'] = "='Net OPEB'!D27"
    rsi[f'{new_col}12'] = "='Net OPEB'!D12"
    rsi[f'{new_col}14'] = "='Net OPEB'!D29"
    rsi[f'{new_col}17'] = inputs.covered_payroll
    rsi[f'{new_col}20'] = f"={new_col}14/{new_col}17"
    rsi[f'{new_col}26'] = "=AmortDeferredOutsIns!C6"
    rsi[f'{new_col}27'] = "Pub-2010/2021"
    rsi[f'{new_col}28'] = "Getzen model"
    
    summary['steps_completed'].append('RSI formulas set')
    summary['rsi_new_column'] = new_col
    
    # =========================================================================
    # STEP 3: SHIFT ARSL VALUES
    # =========================================================================
    amort = wb['AmortDeferredOutsIns']
    amort_data = wb_data['AmortDeferredOutsIns']
    
    # Capture current values
    c13_val = amort_data['C13'].value
    c14_val = amort_data['C14'].value
    c15_val = amort_data['C15'].value
    c16_val = amort_data['C16'].value
    c17_val = amort_data['C17'].value
    c18_val = amort_data['C18'].value
    
    # Shift down
    amort['C14'] = c13_val
    amort['C15'] = c14_val
    amort['C16'] = c15_val
    amort['C17'] = c16_val
    amort['C18'] = c17_val
    amort['C19'] = c18_val
    
    summary['steps_completed'].append('ARSL shifted')
    summary['arsl_shifted'] = {
        'C14': c13_val, 'C15': c14_val, 'C16': c15_val,
        'C17': c16_val, 'C18': c17_val, 'C19': c18_val
    }
    
    # =========================================================================
    # STEP 4: UPDATE A13
    # =========================================================================
    amort['A13'] = new_year
    summary['steps_completed'].append('A13 updated')
    
    # =========================================================================
    # STEP 5: UPDATE ASSUMPTIONS
    # =========================================================================
    assumptions = wb['Assumptions']
    assumptions['C2'] = inputs.valuation_date
    assumptions['C3'] = inputs.prior_measurement_date
    assumptions['C4'] = inputs.measurement_date
    assumptions['C11'] = inputs.prior_discount_rate
    assumptions['C12'] = f"{inputs.new_discount_rate*100:.2f}% annually which is the Bond Buyer 20-Bond General Obligation Index on the Measurement Date.  The 20-Bond Index consists of 20 general obligation bonds that mature in 20 years."
    
    summary['steps_completed'].append('Assumptions updated')
    
    # =========================================================================
    # STEP 6: UPDATE PROVAL1
    # =========================================================================
    proval = wb['ProVal1']
    proval['B19'] = inputs.tol_boy_old_rate
    proval['C19'] = inputs.tol_boy_new_rate
    proval['D19'] = inputs.tol_eoy_baseline
    proval['E19'] = inputs.tol_eoy_disc_plus_1
    proval['F19'] = inputs.tol_eoy_disc_minus_1
    proval['G19'] = inputs.tol_eoy_trend_baseline
    proval['H19'] = inputs.tol_eoy_trend_plus_1
    proval['I19'] = inputs.tol_eoy_trend_minus_1
    proval['B38'] = inputs.service_cost
    proval['D17'] = inputs.covered_payroll
    proval['D88'] = inputs.new_discount_rate
    
    if inputs.active_count is not None:
        proval['D6'] = inputs.active_count
    if inputs.retiree_count is not None:
        proval['D8'] = inputs.retiree_count
    
    summary['steps_completed'].append('ProVal1 updated')
    
    # =========================================================================
    # STEP 7: UPDATE NET OPEB LABELS
    # =========================================================================
    net_opeb = wb['Net OPEB']
    md = inputs.measurement_date
    pmd = inputs.prior_measurement_date
    net_opeb['A9'] = f"Table 3: Changes in Net OPEB Liability for the plan's fiscal year ending {md.month}/{md.day}/{md.year}"
    net_opeb['A12'] = f"Balances at {pmd.month}/{pmd.day}/{pmd.year}"
    net_opeb['A29'] = f"Balances at {md.month}/{md.day}/{md.year}"
    
    summary['steps_completed'].append('Net OPEB labels updated')
    
    # =========================================================================
    # SAVE
    # =========================================================================
    wb.save(output_file)
    wb.close()
    wb_data.close()
    
    summary['output_file'] = output_file
    
    return summary


# =============================================================================
# CONVENIENCE FUNCTION
# =============================================================================

def update_gasb75_rollforward(
    input_file: str,
    output_file: str,
    new_discount_rate: float,
    measurement_date: date,
    covered_payroll: Optional[float] = None,
) -> Dict[str, Any]:
    """
    Convenience function to perform a complete roll-forward update.
    
    Extracts prior year data, calculates roll-forward values, and updates the file.
    
    Args:
        input_file: Prior year GASB 75 file
        output_file: Output file path
        new_discount_rate: New (EOY) discount rate
        measurement_date: New measurement date
        covered_payroll: New covered payroll (optional, uses prior if not provided)
    
    Returns:
        Dict with summary
    """
    # Extract prior year data
    prior = extract_prior_year_data(input_file)
    
    # Calculate roll-forward
    calc = calculate_rollforward(
        prior_tol_eoy=prior.tol_eoy,
        prior_service_cost=prior.service_cost,
        prior_discount_rate=prior.discount_rate,
        new_discount_rate=new_discount_rate
    )
    
    # Determine dates
    prior_measurement_date = date(measurement_date.year - 1, measurement_date.month, measurement_date.day)
    valuation_date = date(measurement_date.year - 1, measurement_date.month + 1, 1) if measurement_date.month < 12 else date(measurement_date.year, 1, 1)
    
    # Create inputs
    inputs = GASB75UpdateInputs(
        valuation_date=valuation_date,
        prior_measurement_date=prior_measurement_date,
        measurement_date=measurement_date,
        prior_discount_rate=prior.discount_rate,
        new_discount_rate=new_discount_rate,
        tol_boy_old_rate=calc['tol_boy_old_rate'],
        tol_boy_new_rate=calc['tol_boy_new_rate'],
        tol_eoy_baseline=calc['tol_eoy_baseline'],
        tol_eoy_disc_plus_1=calc['tol_eoy_disc_plus_1'],
        tol_eoy_disc_minus_1=calc['tol_eoy_disc_minus_1'],
        tol_eoy_trend_baseline=calc['tol_eoy_trend_baseline'],
        tol_eoy_trend_plus_1=calc['tol_eoy_trend_plus_1'],
        tol_eoy_trend_minus_1=calc['tol_eoy_trend_minus_1'],
        service_cost=calc['service_cost'],
        covered_payroll=covered_payroll or prior.covered_payroll,
        active_count=prior.active_count,
        retiree_count=prior.retiree_count,
    )
    
    # Update file
    summary = update_gasb75_file(input_file, output_file, inputs)
    summary['calculation'] = calc
    
    return summary


# =============================================================================
# EXAMPLE / TEST
# =============================================================================

if __name__ == '__main__':
    print("GASB 75 Updater Module")
    print("=" * 50)
    print()
    print("This module provides:")
    print("  - extract_prior_year_data(file) - Extract data from prior file")
    print("  - calculate_rollforward(...) - Calculate roll-forward values")
    print("  - update_gasb75_file(input, output, inputs) - Update file")
    print("  - update_gasb75_rollforward(...) - Convenience roll-forward")
    print()
    print("Example usage:")
    print()
    print("  from gasb75_updater import update_gasb75_rollforward")
    print("  from datetime import date")
    print()
    print("  summary = update_gasb75_rollforward(")
    print("      input_file='GASB75_2024.xlsx',")
    print("      output_file='GASB75_2025.xlsx',")
    print("      new_discount_rate=0.0502,")
    print("      measurement_date=date(2025, 9, 30)")
    print("  )")
    print()
    print("See GASB75_UPDATE_INSTRUCTIONS.md for detailed documentation.")
